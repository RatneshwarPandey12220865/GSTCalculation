import io
import re
import urllib.request
import wave
from collections import Counter

import ddddocr
import numpy as np
import openpyxl
import streamlit as st
from PIL import Image, ImageEnhance, ImageFilter
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth


# ── CAPTCHA: Image preprocessing + OCR ───────────────────────────────────────

_OCR = None
_OCR_OLD = None


def _get_ocr():
    global _OCR
    if _OCR is None:
        _OCR = ddddocr.DdddOcr(show_ad=False)
    return _OCR


def _get_ocr_old():
    """Old model is significantly more accurate for plain digit CAPTCHAs."""
    global _OCR_OLD
    if _OCR_OLD is None:
        _OCR_OLD = ddddocr.DdddOcr(old=True, show_ad=False)
    return _OCR_OLD


def _img_to_bytes(img: Image.Image) -> bytes:
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def _keep_dark(img_rgb: Image.Image, thresh: int) -> Image.Image:
    arr = np.array(img_rgb, dtype=np.int32)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    dark = (r < thresh) & (g < thresh) & (b < thresh)
    out = np.full(arr.shape[:2], 255, dtype=np.uint8)
    out[dark] = 0
    return Image.fromarray(out, "L")


def _remove_red_and_grid(img_rgb: Image.Image) -> Image.Image:
    arr = np.array(img_rgb, dtype=np.int32)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    arr[(r > 120) & (r > g + 45) & (r > b + 45)] = [255, 255, 255]  # red line
    arr[(r > 170) & (g > 170) & (b > 170)] = [255, 255, 255]         # grid
    return Image.fromarray(arr.astype(np.uint8), "RGB")


def _close_gaps(img_l: Image.Image, size: int = 3) -> Image.Image:
    return img_l.filter(ImageFilter.MinFilter(size)).filter(ImageFilter.MaxFilter(size))


def _captcha_variants(img_bytes: bytes) -> list[bytes]:
    base = Image.open(io.BytesIO(img_bytes)).convert("RGB")

    v_loose  = _keep_dark(base, 115)
    v_tight  = _keep_dark(base, 95)
    v_very_tight = _keep_dark(base, 80)
    v_closed = _close_gaps(v_loose, 3)

    cleaned = _remove_red_and_grid(base).convert("L")
    cleaned = ImageEnhance.Contrast(cleaned).enhance(2.5)
    cleaned = cleaned.filter(ImageFilter.SHARPEN)
    v_clean = cleaned.point(lambda p: 0 if p < 135 else 255, "L")

    # Extra variant: aggressive denoising then threshold
    cleaned2 = _remove_red_and_grid(base).convert("L")
    cleaned2 = ImageEnhance.Contrast(cleaned2).enhance(3.0)
    cleaned2 = cleaned2.filter(ImageFilter.MedianFilter(3))
    v_clean2 = cleaned2.point(lambda p: 0 if p < 120 else 255, "L")

    out = []
    for v in [v_loose, v_tight, v_very_tight, v_closed, v_clean, v_clean2]:
        w, h = v.size
        out.append(_img_to_bytes(v.resize((w * 2, h * 2), Image.LANCZOS)))
    return out


def _ocr_classify(ocr_instance, img_bytes: bytes) -> str:
    try:
        return ocr_instance.classification(img_bytes).strip()
    except Exception:
        return ""


# ── TrOCR (HuggingFace printed-text OCR) ─────────────────────────────────────

_TROCR_PROCESSOR = None
_TROCR_MODEL = None


def _get_trocr():
    """
    Lazy-load microsoft/trocr-small-printed (~346 MB, downloaded once).
    Better than ddddocr on noisy/distorted printed digits.
    """
    global _TROCR_PROCESSOR, _TROCR_MODEL
    if _TROCR_PROCESSOR is None:
        from transformers import TrOCRProcessor, VisionEncoderDecoderModel
        _TROCR_PROCESSOR = TrOCRProcessor.from_pretrained("microsoft/trocr-small-printed")
        _TROCR_MODEL     = VisionEncoderDecoderModel.from_pretrained("microsoft/trocr-small-printed")
        _TROCR_MODEL.eval()
        print("[trocr] Model loaded OK")
    return _TROCR_PROCESSOR, _TROCR_MODEL


def _trocr_classify(img_bytes: bytes) -> str:
    try:
        import torch
        processor, model = _get_trocr()
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        pixel_values = processor(images=img, return_tensors="pt").pixel_values
        with torch.no_grad():
            generated_ids = model.generate(pixel_values, max_new_tokens=20)
        text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0].strip()
        return text
    except Exception as e:
        print(f"[trocr] classify failed: {e}")
        return ""


def solve_captcha_image(img_bytes: bytes) -> str | None:
    ocr     = _get_ocr()
    ocr_old = _get_ocr_old()

    try:
        variants = _captcha_variants(img_bytes)
    except Exception as e:
        print(f"[ocr] Variant generation failed: {e}")
        variants = []

    candidates: list[str] = []

    # ddddocr (new + old) on every variant + raw image
    for img in variants + [img_bytes]:
        for model_label, model in [("new", ocr), ("old", ocr_old)]:
            raw = _ocr_classify(model, img)
            if not raw:
                continue
            digits = re.sub(r"[^0-9]", "", raw)
            safe_raw = raw.encode("ascii", errors="replace").decode()
            print(f"[ocr/{model_label}] raw='{safe_raw}' -> digits='{digits}'")
            if len(digits) == 6:
                candidates.append(digits)

    # TrOCR on raw image + first 3 variants (skip rest to stay fast)
    for img in ([img_bytes] + variants[:3]):
        raw = _trocr_classify(img)
        if not raw:
            continue
        digits = re.sub(r"[^0-9]", "", raw)
        safe_raw = raw.encode("ascii", errors="replace").decode()
        print(f"[ocr/trocr] raw='{safe_raw}' -> digits='{digits}'")
        if len(digits) == 6:
            candidates.append(digits)

    if not candidates:
        print("[ocr] No variant produced 6 digits -- will fall back to audio.")
        return None

    best, count = Counter(candidates).most_common(1)[0]
    print(f"[ocr] candidates={candidates} -> chose '{best}' ({count}/{len(candidates)} agree)")
    return best


# ── CAPTCHA: Audio fallback (HuggingFace Whisper, offline after first download) ─

_WORD_TO_DIGIT = {
    "zero": "0", "oh": "0", "o": "0", "one": "1", "won": "1", "two": "2",
    "to": "2", "too": "2", "tu": "2", "three": "3", "tree": "3", "free": "3",
    "four": "4", "for": "4", "fore": "4", "five": "5", "fife": "5", "fire": "5",
    "six": "6", "sic": "6", "seven": "7", "eight": "8", "ate": "8", "ait": "8",
    "nine": "9", "niner": "9", "nein": "9",
}

_WHISPER_PIPE = None


def _get_whisper():
    """
    Lazy-load the HuggingFace Whisper pipeline (singleton).
    Uses openai/whisper-base (~74 MB) for better digit accuracy than whisper-tiny.
    Downloaded once and cached by HuggingFace in ~/.cache/huggingface.
    """
    global _WHISPER_PIPE
    if _WHISPER_PIPE is None:
        from transformers import pipeline
        _WHISPER_PIPE = pipeline(
            "automatic-speech-recognition",
            model="openai/whisper-tiny",
            max_new_tokens=30,          # 6 digits need ≤15 tokens; cap prevents runaway on silence
            generate_kwargs={"language": "english", "task": "transcribe"},
        )
    return _WHISPER_PIPE


def _decode_audio(audio_data: bytes) -> tuple[np.ndarray, int]:
    """
    Decode audio bytes to float32 mono at 16 kHz.
    Tries soundfile first (handles WAV + OGG + FLAC), falls back to stdlib WAV.
    """
    try:
        import soundfile as sf  # type: ignore
        samples, sr = sf.read(io.BytesIO(audio_data), dtype="float32", always_2d=False)
        if samples.ndim == 2:
            samples = samples.mean(axis=1)
        if sr != 16000:
            target_len = int(len(samples) * 16000 / sr)
            samples = np.interp(
                np.linspace(0, len(samples) - 1, target_len),
                np.arange(len(samples)),
                samples,
            ).astype(np.float32)
        print(f"[audio] soundfile decoded: {len(samples)} samples at 16kHz")
        return samples.astype(np.float32), 16000
    except Exception as e:
        print(f"[audio] soundfile failed ({e}), trying stdlib WAV decoder")

    # Stdlib fallback — WAV only
    with wave.open(io.BytesIO(audio_data), "rb") as wf:
        n_channels = wf.getnchannels()
        sampwidth  = wf.getsampwidth()
        framerate  = wf.getframerate()
        frames     = wf.readframes(wf.getnframes())

    if sampwidth == 1:
        samples = (np.frombuffer(frames, dtype=np.uint8).astype(np.float32) - 128) / 128.0
    elif sampwidth == 2:
        samples = np.frombuffer(frames, dtype=np.int16).astype(np.float32) / 32768.0
    elif sampwidth == 4:
        samples = np.frombuffer(frames, dtype=np.int32).astype(np.float32) / 2147483648.0
    else:
        raise ValueError(f"Unsupported WAV sample width: {sampwidth} bytes")

    if n_channels > 1:
        samples = samples.reshape(-1, n_channels).mean(axis=1)

    if framerate != 16000:
        target_len = int(len(samples) * 16000 / framerate)
        samples = np.interp(
            np.linspace(0, len(samples) - 1, target_len),
            np.arange(len(samples)),
            samples,
        ).astype(np.float32)

    return samples, 16000


def _words_to_digits(text: str) -> str:
    result = []
    # Remove punctuation and split
    tokens = re.sub(r"[^a-z0-9\s]", " ", text.lower()).split()
    for token in tokens:
        if token in _WORD_TO_DIGIT:
            result.append(_WORD_TO_DIGIT[token])
        else:
            # keep any bare digit characters in the token
            digits_in_token = re.sub(r"[^0-9]", "", token)
            if digits_in_token:
                result.append(digits_in_token)
    return "".join(result)


def _transcribe_audio(audio_data: bytes) -> str | None:
    try:
        samples, rate = _decode_audio(audio_data)
        pipe = _get_whisper()
        result = pipe({"array": samples, "sampling_rate": rate})
        text = (result.get("text") or "").strip()
        print(f"[audio] Whisper raw output: '{text}'")
        return text or None
    except Exception as e:
        print(f"[audio] Transcription failed: {e}")
        return None


def _is_audio_url(url: str) -> bool:
    u = url.lower()
    return any(kw in u for kw in [
        ".mp3", ".wav", ".ogg", ".m4a", ".aac",
        "audio", "sound", "captcha", "kaptcha", "verifycode",
    ])


def solve_captcha_audio(page) -> str | None:
    # Capture audio response BODY directly from the browser's network traffic
    # so we don't need to re-download without session cookies.
    captured_audio: list[bytes] = []
    captured_url:   list[str]   = []

    def handle_response(response):
        if captured_audio:
            return  # already got one
        if not _is_audio_url(response.url):
            return
        try:
            body = response.body()
            if body and len(body) > 512:  # skip tiny/empty responses
                captured_audio.append(body)
                captured_url.append(response.url)
                print(f"[audio] Captured {len(body)} bytes from {response.url}")
        except Exception as e:
            print(f"[audio] Could not read response body: {e}")
            captured_url.append(response.url)

    page.on("response", handle_response)

    # Locate the sound/audio button
    sound_btn = None
    for btn in page.get_by_role("button").all():
        try:
            aria  = (btn.get_attribute("aria-label") or "").lower()
            title = (btn.get_attribute("title") or "").lower()
            txt   = btn.inner_text().strip().lower()
            if any(kw in aria + title + txt for kw in ["sound", "audio", "speaker", "listen"]):
                sound_btn = btn
                break
        except Exception:
            pass
    if sound_btn is None:
        sound_btn = page.get_by_role("button").nth(4)

    try:
        sound_btn.click()
    except Exception as e:
        print(f"[audio] Could not click sound button: {e}")
        page.remove_listener("response", handle_response)
        return None

    print("[audio] Clicked sound button. Waiting 12 seconds for audio...")
    page.wait_for_timeout(12000)

    try:
        page.remove_listener("response", handle_response)
    except Exception:
        pass

    # Case 1: we captured the body directly
    if captured_audio:
        audio_data = captured_audio[0]
    elif captured_url:
        # Case 2: captured URL but body read failed — try DOM then give up
        audio_data = None
    else:
        # Case 3: no network hit — try DOM <audio> element
        try:
            dom_src = page.evaluate(
                """() => {
                    const el = document.querySelector('audio source, audio');
                    return el ? (el.src || el.getAttribute('src') || '') : '';
                }"""
            ) or ""
            if dom_src and _is_audio_url(dom_src):
                captured_url.append(dom_src)
        except Exception:
            pass
        audio_data = None

    # If we have a URL but no body, fetch via Playwright's own context (has session)
    if audio_data is None and captured_url:
        url = captured_url[0]
        print(f"[audio] Fetching audio via Playwright request: {url}")
        try:
            resp = page.request.get(url)
            audio_data = resp.body()
            print(f"[audio] Fetched {len(audio_data)} bytes")
        except Exception as e:
            print(f"[audio] Playwright fetch failed: {e}")

    if not audio_data:
        print("[audio] No audio data obtained.")
        return None

    text = _transcribe_audio(audio_data)
    if not text:
        print("[audio] Transcription produced no text.")
        return None
    print(f"[audio] Transcription: '{text}'")

    digits = _words_to_digits(text)
    print(f"[audio] Digits extracted: '{digits}'")
    if len(digits) == 6:
        return digits
    print(f"[audio] Expected 6 digits, got {len(digits)} — discarding.")
    return None


def refresh_captcha(page):
    page.get_by_role("cell").get_by_role("button").click()
    page.wait_for_timeout(1500)
    print("[captcha] Refreshed.")


# ── Detail page extraction ────────────────────────────────────────────────────

def extract_detail_page(page, gstin: str) -> dict:
    page.wait_for_selector("text=Search Result based on GSTIN", timeout=15000)
    page.wait_for_timeout(1500)

    data = page.evaluate("""() => {
        const out = {};
        const txt = el => (el ? el.innerText.trim() : "");

        document.querySelectorAll('#lottable .col-sm-6').forEach(col => {
            const bold = col.querySelector('b');
            if (!bold) return;
            const label = txt(bold).replace(/:$/, '').trim();
            if (!label) return;
            const clone = col.cloneNode(true);
            const boldClone = clone.querySelector('b');
            if (boldClone) boldClone.remove();
            const value = clone.innerText.trim();
            if (value) out[label] = value;
        });

        if (Object.keys(out).length < 3) {
            document.querySelectorAll('b, strong').forEach(bold => {
                const label = txt(bold).replace(/:$/, '').trim();
                if (!label || label.length > 60) return;
                let next = bold.parentElement && bold.parentElement.nextElementSibling;
                if (!next) next = bold.nextElementSibling;
                if (next) out[label] = txt(next);
            });
        }

        document.querySelectorAll('.panel').forEach(panel => {
            const heading = txt(panel.querySelector('.panel-heading, .panel-title') || panel.querySelector('[class*="heading"]'));
            const body    = panel.querySelector('.panel-body, [class*="body"]');
            if (!heading || !body) return;
            const h = heading.toLowerCase();
            if (h.includes('nature of core business')) {
                const link = body.querySelector('a');
                out['Nature Of Core Business Activity'] = link ? txt(link) : txt(body);
            }
            if (h.includes('nature of business activities')) {
                const items = body.querySelectorAll('li');
                if (items.length > 0) {
                    out['Nature of Business Activities'] = Array.from(items)
                        .map(li => txt(li).replace(/^\\d+\\.\\s*/, ''))
                        .join('\\n');
                } else {
                    out['Nature of Business Activities'] = txt(body);
                }
            }
        });

        const goodsHSN = [], goodsDesc = [], servicesHSN = [], servicesDesc = [];
        document.querySelectorAll('table').forEach(table => {
            const allText = txt(table).toLowerCase();
            if (!allText.includes('goods') && !allText.includes('hsn')) return;
            Array.from(table.querySelectorAll('tr')).forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length === 0) return;
                const c0 = txt(cells[0]);
                if (!c0 || c0.toLowerCase() === 'hsn' || c0.toLowerCase() === 'goods'
                         || c0.toLowerCase() === 'services') return;
                const c1 = cells[1] ? txt(cells[1]) : '';
                const c2 = cells[2] ? txt(cells[2]) : '';
                const c3 = cells[3] ? txt(cells[3]) : '';
                if (c0) { goodsHSN.push(c0);  goodsDesc.push(c1); }
                if (c2) { servicesHSN.push(c2); servicesDesc.push(c3); }
            });
        });
        out['Goods HSN']         = goodsHSN.join('\\n');
        out['Goods Description'] = goodsDesc.join('\\n');
        out['Services HSN']      = servicesHSN.join('\\n');
        out['Services Desc']     = servicesDesc.join('\\n');

        out['__debug_keys__'] = Object.keys(out).filter(k => !k.startsWith('_')).join(' | ');
        return out;
    }""")

    print(f"[extract] Raw keys from JS: {(data or {}).get('__debug_keys__', 'NONE')}")

    def get(key: str) -> str:
        for k, v in (data or {}).items():
            if k.strip().lower() == key.strip().lower():
                return (v or "").strip()
        return ""

    return {
        "gstin":                            gstin,
        "legal_name":                       get("Legal Name of Business"),
        "trade_name":                       get("Trade Name"),
        "effective_date_of_reg":            get("Effective Date of registration"),
        "constitution_of_business":         get("Constitution of Business"),
        "gstin_uin_status":                 get("GSTIN / UIN Status"),
        "taxpayer_type":                    get("Taxpayer Type"),
        "administrative_office":            get("Administrative Office"),
        "other_office":                     get("Other Office"),
        "principal_place_of_business":      get("Principal Place of Business"),
        "aadhaar_authenticated":            get("Whether Aadhaar Authenticated?"),
        "ekyc_verified":                    get("Whether e-KYC Verified?"),
        "additional_trade_name":            get("Additional Trade Name"),
        "nature_of_core_business":          get("Nature Of Core Business Activity"),
        "nature_of_business_activities":    get("Nature of Business Activities"),
        "goods_hsn":                        get("Goods HSN"),
        "goods_description":                get("Goods Description"),
        "services_hsn":                     get("Services HSN"),
        "services_description":             get("Services Desc"),
    }


# ── Core scraping ─────────────────────────────────────────────────────────────

def _solve_and_submit(page, max_attempts: int = 8) -> bool:
    """
    For each attempt:
      1. Try OCR once on the current captcha image.
      2. If OCR fails → immediately try audio (no further OCR refreshes).
      3. If both fail → refresh captcha and try again.
      4. If either succeeds → submit and check result.
    """
    page.wait_for_selector("#imgCaptcha", timeout=15000)
    page.wait_for_timeout(1000)

    for attempt in range(max_attempts):
        print(f"\n[captcha] Attempt {attempt + 1}/{max_attempts}")

        # Step 1: single OCR attempt
        captcha_bytes = page.locator("#imgCaptcha").screenshot()
        captcha_text = solve_captcha_image(captcha_bytes)

        if captcha_text:
            print(f"[captcha] OCR succeeded: '{captcha_text}'")
        else:
            # Step 2: OCR failed -> go straight to audio
            print("[captcha] OCR failed -- trying audio immediately...")
            captcha_text = solve_captcha_audio(page)
            if captcha_text:
                print(f"[captcha] Audio succeeded: '{captcha_text}'")
            else:
                print("[captcha] Audio also failed -- refreshing captcha and retrying.")
                refresh_captcha(page)
                continue

        # Step 3: submit whatever we got
        print(f"[captcha] Submitting: '{captcha_text}'")
        captcha_box = page.get_by_role("textbox", name="Type the characters you see")
        captcha_box.clear()
        captcha_box.fill(captcha_text)
        page.get_by_role("button", name="Search", exact=True).click()
        page.wait_for_timeout(3000)

        if page.get_by_text("Enter valid letters shown in").is_visible():
            print("[captcha] Wrong captcha -- refreshing and retrying.")
            refresh_captcha(page)
            continue

        return True

    return False


def scrape_gst_single(gstin: str, page) -> dict | None:
    gstin_box = page.get_by_role("textbox", name="Enter GSTIN/UIN of the")
    gstin_box.wait_for(timeout=15000)
    gstin_box.fill(gstin)
    gstin_box.press("Enter")

    if not _solve_and_submit(page):
        print(f"[scrape] Could not solve captcha for {gstin}")
        return None

    try:
        return extract_detail_page(page, gstin)
    except Exception as e:
        print(f"[scrape] Extraction error for {gstin}: {e}")
        return None


def _browser_context(playwright):
    browser = playwright.chromium.launch(
        headless=False,
        args=["--disable-blink-features=AutomationControlled", "--start-maximized"]
    )
    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="en-IN",
        timezone_id="Asia/Kolkata",
        viewport={"width": 1280, "height": 800},
        extra_http_headers={"Accept-Language": "en-IN,en;q=0.9"},
    )
    page = context.new_page()
    Stealth().apply_stealth_sync(page)
    return browser, context, page


def scrape_gst(gstin: str) -> dict | None:
    with sync_playwright() as p:
        browser, context, page = _browser_context(p)
        try:
            page.goto("https://services.gst.gov.in/services/searchtp", timeout=60000)
            page.wait_for_load_state("networkidle")
            return scrape_gst_single(gstin, page)
        finally:
            context.close()
            browser.close()


def scrape_gst_batch(gstins: list[str], progress_cb=None) -> list[dict]:
    results = []
    with sync_playwright() as p:
        browser, context, page = _browser_context(p)
        try:
            page.goto("https://services.gst.gov.in/services/searchtp", timeout=60000)
            page.wait_for_load_state("networkidle")

            for i, gstin in enumerate(gstins):
                gstin = gstin.strip().upper()
                if not gstin:
                    continue
                print(f"\n{'='*50}\n[batch] Processing: {gstin}\n{'='*50}")
                if progress_cb:
                    progress_cb(i, len(gstins), gstin)
                data = scrape_gst_single(gstin, page)
                results.append(data if data else {"gstin": gstin, "error": "Failed to extract"})

                try:
                    page.get_by_role("textbox", name="Enter GSTIN/UIN of the").wait_for(timeout=5000)
                except Exception:
                    page.goto("https://services.gst.gov.in/services/searchtp", timeout=60000)
                    page.wait_for_load_state("networkidle")
        finally:
            context.close()
            browser.close()

    return results


# ── Excel helpers ─────────────────────────────────────────────────────────────

EXCEL_COLUMNS = [
    ("gstin",                           "GSTIN"),
    ("legal_name",                      "Legal Name of Business"),
    ("trade_name",                      "Trade Name"),
    ("effective_date_of_reg",           "Effective Date of Registration"),
    ("constitution_of_business",        "Constitution of Business"),
    ("gstin_uin_status",                "GSTIN / UIN Status"),
    ("taxpayer_type",                   "Taxpayer Type"),
    ("administrative_office",           "Administrative Office"),
    ("other_office",                    "Other Office"),
    ("principal_place_of_business",     "Principal Place of Business"),
    ("aadhaar_authenticated",           "Aadhaar Authenticated?"),
    ("ekyc_verified",                   "e-KYC Verified?"),
    ("additional_trade_name",           "Additional Trade Name"),
    ("nature_of_core_business",         "Nature Of Core Business Activity"),
    ("nature_of_business_activities",   "Nature of Business Activities"),
    ("goods_hsn",                       "Goods HSN"),
    ("goods_description",               "Goods Description"),
    ("services_hsn",                    "Services HSN"),
    ("services_description",            "Services Description"),
    ("error",                           "Error"),
]


def results_to_excel(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Results"
    ws.append([col[1] for col in EXCEL_COLUMNS])
    for row in results:
        ws.append([row.get(col[0], "") for col in EXCEL_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def read_gstins_from_excel(data: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    header_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().upper() == "GSTNO":
            header_col = col_idx
            break
    if header_col is None:
        raise ValueError("Column header 'GSTNO' not found in the uploaded Excel file.")
    gstins = []
    for row in ws.iter_rows(min_row=2, min_col=header_col, max_col=header_col, values_only=True):
        val = row[0]
        if val and str(val).strip():
            gstins.append(str(val).strip().upper())
    return gstins


# ── Streamlit UI ──────────────────────────────────────────────────────────────

DISPLAY_FIELDS = [
    ("legal_name",                   "Legal Name of Business"),
    ("trade_name",                   "Trade Name"),
    ("effective_date_of_reg",        "Effective Date of Registration"),
    ("constitution_of_business",     "Constitution of Business"),
    ("gstin_uin_status",             "GSTIN / UIN Status"),
    ("taxpayer_type",                "Taxpayer Type"),
    ("principal_place_of_business",  "Principal Place of Business"),
    ("administrative_office",        "Administrative Office"),
    ("other_office",                 "Other Office"),
    ("aadhaar_authenticated",        "Aadhaar Authenticated?"),
    ("ekyc_verified",                "e-KYC Verified?"),
    ("additional_trade_name",        "Additional Trade Name"),
    ("nature_of_core_business",      "Nature Of Core Business Activity"),
    ("nature_of_business_activities","Nature of Business Activities"),
    ("goods_hsn",                    "Goods HSN Codes"),
    ("goods_description",            "Goods Description"),
    ("services_hsn",                 "Services HSN Codes"),
    ("services_description",         "Services Description"),
]


def _render_result(result: dict):
    status = result.get("gstin_uin_status", "")
    col1, col2 = st.columns([3, 1])
    with col1:
        st.subheader(result.get("gstin", ""))
    with col2:
        if "active" in status.lower() and "inactive" not in status.lower():
            st.success(status)
        elif "inactive" in status.lower():
            st.error(status)
        else:
            st.info(status or "—")

    for key, label in DISPLAY_FIELDS:
        val = result.get(key, "") or ""
        if val:
            st.markdown(f"**{label}**")
            st.write(val)
            st.divider()


def main():
    st.set_page_config(page_title="GST Taxpayer Search", page_icon="⚡", layout="centered")
    st.title("GST Taxpayer Search")
    st.caption("CAPTCHA is solved automatically. Powered by OCR + Offline Audio.")

    tab_single, tab_batch = st.tabs(["Single GSTIN", "Batch Upload (Excel)"])

    # ── Single GSTIN tab ──────────────────────────────────────────────────────
    with tab_single:
        gstin_input = st.text_input(
            "Enter GSTIN / UIN",
            placeholder="e.g. 06AJWPP4177G1ZL",
            max_chars=15,
        )

        if st.button("Search", type="primary", key="btn_single"):
            gstin = gstin_input.strip().upper()
            if not gstin:
                st.warning("Please enter a GSTIN.")
            elif len(gstin) != 15:
                st.warning("GSTIN must be exactly 15 characters.")
            else:
                with st.spinner("Opening browser & solving CAPTCHA automatically..."):
                    try:
                        result = scrape_gst(gstin)
                    except Exception as e:
                        result = None
                        st.error(f"Error: {e}")

                if result:
                    st.success("Data fetched successfully!")
                    _render_result(result)
                elif result is not None:
                    st.error("No data found for this GSTIN.")

    # ── Batch tab ─────────────────────────────────────────────────────────────
    with tab_batch:
        st.markdown(
            "Upload an Excel file with a column header named **GSTNO**. "
            "All rows below it will be processed one by one."
        )
        uploaded = st.file_uploader("Choose Excel file", type=["xlsx", "xls"])

        if uploaded and st.button("Process Batch", type="primary", key="btn_batch"):
            try:
                gstins = read_gstins_from_excel(uploaded.read())
            except ValueError as e:
                st.error(str(e))
                return

            if not gstins:
                st.error("No GSTINs found under the 'GSTNO' column.")
                return

            st.info(f"Found {len(gstins)} GSTIN(s). Starting browser...")
            progress_bar  = st.progress(0)
            status_text   = st.empty()

            def on_progress(i, total, current_gstin):
                progress_bar.progress(i / total)
                status_text.text(f"Processing {i + 1}/{total}: {current_gstin}")

            with st.spinner("Processing... this may take a few minutes."):
                try:
                    results = scrape_gst_batch(gstins, progress_cb=on_progress)
                except Exception as e:
                    st.error(f"Batch error: {e}")
                    return

            progress_bar.progress(1.0)
            status_text.text("Done!")
            st.success(f"All {len(results)} GSTIN(s) processed.")

            excel_bytes = results_to_excel(results)
            st.download_button(
                label="Download Result Excel",
                data=excel_bytes,
                file_name="gst_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # Preview results in a table
            import pandas as pd
            df = pd.DataFrame(results)
            st.dataframe(df, use_container_width=True)


if __name__ == "__main__":
    main()
